# -*- coding: utf-8 -*-
"""
메타_시간대분석.py  (읽기 전용 — Supabase/시트 미수정)
======================================================
Meta Marketing API 로 전체 광고 데이터를 받아 **상품별 × 시간대(0~23시)별**
판매수(구매수)와 구매당비용(CPA)을 집계한다.

기간: 2026-05-01 ~ 2026-06-08 (고정, --since/--until 로 override 가능)
시간대: Meta breakdowns=hourly_stats_aggregated_by_advertiser_time_zone (광고주 TZ 기준 0~23시)
상품: 캠페인/광고세트 이름에서 추출 (국내_세트별_supabase.py 의 extract_product 와 동일 규칙)
판매수: actions 중 purchase 류 합
구매당비용(CPA): spend / 판매수  (통화는 계정 통화 — 국내 KRW, 글로벌 USD)

환경변수(토큰):
  META_TOKEN_1   (국내 act_1270614404675034/act_707835224206178, 글로벌 act_1054081590008088)
  META_TOKEN_2   (국내 act_1808141386564262)
  META_TOKEN_GlobalTT / META_TOKEN_4 / META_TOKEN_3  (글로벌 act_2677.../act_1335...)
  META_TOKEN_ACT_9937  (대만 act_993712016404855)

옵션:
  --scope kr|global|all   (기본 all — 토큰 있는 계정만 자동 시도)
  --since YYYY-MM-DD --until YYYY-MM-DD
  --level adset|campaign  (기본 adset)
"""
import os, re, sys, time, json
from collections import defaultdict
from pathlib import Path
import requests

# 로컬 실행 시 ./.env 또는 ../meta_scraper/.env 자동 로드 (토큰/키)
def _load_env():
    for p in [Path(__file__).parent / ".env",
              Path(__file__).parent.parent / "meta_scraper" / ".env"]:
        if p.exists():
            for line in p.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    val = v.strip().strip('"').strip("'")
                    if val:   # 빈 값은 스킵 — 다른 .env 의 실제 토큰을 빈 값으로 덮지 않도록
                        os.environ.setdefault(k.strip(), val)
_load_env()

API = "v21.0"
BASE = f"https://graph.facebook.com/{API}"

# 계정 → (토큰env, 통화, 권역)
ACCOUNTS = [
    ("act_1270614404675034", "META_TOKEN_1",      "KRW", "kr"),
    ("act_707835224206178",  "META_TOKEN_1",      "KRW", "kr"),
    ("act_1808141386564262", "META_TOKEN_2",      "KRW", "kr"),
    ("act_1054081590008088", "META_TOKEN_1",      "USD", "global"),
    ("act_2677707262628563", "META_TOKEN_GlobalTT","USD","global"),
    ("act_1335040608536838", "META_TOKEN_GlobalTT","USD","global"),
    ("act_993712016404855",  "META_TOKEN_ACT_9937","USD","global"),
]

def argval(flag, default=None):
    return sys.argv[sys.argv.index(flag)+1] if flag in sys.argv else default

SCOPE = argval("--scope", "kr")   # 기본 국내만 (글로벌 USD 상품 제외). --scope all 로 전체.
SINCE = argval("--since", "2026-05-01")
UNTIL = argval("--until", "2026-06-08")
LEVEL = argval("--level", "adset")
PURCHASE_TYPES = {"purchase", "omni_purchase", "offsite_conversion.fb_pixel_purchase"}

def token_for(env_name):
    # 계정별 토큰 env 이름이 .env 마다 다를 수 있어 별칭 체인으로 탐색.
    # (없거나 권한 없으면 fetch 시 에러 → 해당 계정만 스킵되므로 fallback 추가는 안전)
    CHAINS = {
        "META_TOKEN_GlobalTT": ["META_TOKEN_GlobalTT", "META_TOKEN_GLOBAL1",
                                 "META_TOKEN_4", "META_TOKEN_3", "META_TOKEN_1"],
        "META_TOKEN_ACT_9937": ["META_TOKEN_ACT_9937", "META_TOKEN_SAJU_TW",
                                 "META_TOKEN_GLOBAL1"],
    }
    for name in CHAINS.get(env_name, [env_name]):
        v = os.environ.get(name)
        if v:
            return v
    return ""

def _strip_leading_emojis(text):
    i = 0
    while i < len(text):
        c = text[i]
        if ("가" <= c <= "힣" or "ㄱ" <= c <= "ㅣ" or c.isalnum() or c in ".%"):
            break
        i += 1
    return text[i:].strip()

def extract_product(adset_name, campaign_name=""):
    for source in [campaign_name, adset_name]:
        if not source: continue
        cleaned = _strip_leading_emojis(str(source).strip())
        if not cleaned: continue
        for token in re.split(r"[_\s\-/|,()\[\]]+", cleaned):
            token = token.strip()
            if not token: continue
            if re.match(r"^\d+$", token): continue
            return token
    return "기타"

def fetch_insights(acc, token):
    url = f"{BASE}/{acc}/insights"
    params = {
        "access_token": token,
        "level": LEVEL,
        "time_range": json.dumps({"since": SINCE, "until": UNTIL}),
        "breakdowns": "hourly_stats_aggregated_by_advertiser_time_zone",
        "fields": "campaign_name,adset_name,spend,actions",
        "limit": "500",
    }
    rows, page = [], 0
    while url:
        r = requests.get(url, params=params if page == 0 else None, timeout=120)
        if r.status_code != 200:
            print(f"  ❌ {acc} HTTP {r.status_code}: {r.text[:200]}")
            break
        j = r.json()
        rows.extend(j.get("data", []))
        nxt = j.get("paging", {}).get("next")
        url = nxt; params = None; page += 1
        if page > 200: break
    return rows

def hour_of(bucket):
    # "00:00:00 - 00:59:59" → 0
    m = re.match(r"^(\d{1,2})", str(bucket or ""))
    return int(m.group(1)) if m else -1

def purchases_of(actions):
    # ⚠️ purchase / omni_purchase / offsite_conversion.fb_pixel_purchase 는 보통
    #    같은 구매를 중복 표기한다(값 동일). 합산하면 2~3배 뻥튀기되어 결과당비용이
    #    1/3로 싸게 나온다. → 첫 매칭 1개만 사용 (국내_세트별_supabase.py 규칙과 동일).
    for a in actions or []:
        if a.get("action_type") in PURCHASE_TYPES:
            try: return float(a.get("value", 0))
            except: return 0.0
    return 0.0

def _desktop_dir():
    home = Path.home()
    for c in [home/"Desktop", home/"OneDrive"/"Desktop",
              home/"OneDrive"/"바탕 화면", home/"바탕 화면"]:
        if c.is_dir():
            return c
    return home

def write_excel(agg, prod_cur, prod_tot, prods, since, until):
    """상품합계 / 상품별×시간대 / 통화별×시간대 3개 시트로 바탕화면에 .xlsx 저장."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A2744")
    tot_fill = PatternFill("solid", fgColor="FFF2CC")
    num = "#,##0"
    center = Alignment(horizontal="center")

    def style_header(ws, ncol):
        for c in range(1, ncol+1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        ws.freeze_panes = "A2"

    def autowidth(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # 1) 상품합계
    ws1 = wb.active; ws1.title = "상품합계"
    ws1.append(["상품", "통화", "총판매수", "총지출", "결과당비용"])
    for p in prods:
        t = prod_tot[p]
        if t["purch"] < 1 and t["spend"] < 1: continue
        cpa = t["spend"]/t["purch"] if t["purch"] > 0 else 0
        ws1.append([p, prod_cur.get(p, ""), round(t["purch"]), round(t["spend"]), round(cpa)])
    for r in range(2, ws1.max_row+1):
        for c in (3, 4, 5): ws1.cell(row=r, column=c).number_format = num
    style_header(ws1, 5); autowidth(ws1, [18, 7, 12, 16, 14])

    # 2) 상품별 × 시간대
    ws2 = wb.create_sheet("상품별_시간대")
    ws2.append(["상품", "통화", "시간대", "판매수", "지출", "결과당비용"])
    for p in prods:
        cur = prod_cur.get(p, "")
        for h in range(24):
            v = agg.get((p, h))
            if not v or (v["purch"] < 1 and v["spend"] < 1): continue
            cpa = v["spend"]/v["purch"] if v["purch"] > 0 else 0
            ws2.append([p, cur, f"{h:02d}시", round(v["purch"]), round(v["spend"]), round(cpa)])
    for r in range(2, ws2.max_row+1):
        for c in (4, 5, 6): ws2.cell(row=r, column=c).number_format = num
    style_header(ws2, 6); autowidth(ws2, [18, 7, 9, 10, 16, 14])

    # 3) 통화별 × 시간대 (+ 통화 합계)
    ws3 = wb.create_sheet("통화별_시간대")
    ws3.append(["통화", "시간대", "판매수", "지출", "결과당비용"])
    by_cur_h = defaultdict(lambda: defaultdict(lambda: {"spend": 0.0, "purch": 0.0}))
    for (p, h), v in agg.items():
        cur = prod_cur.get(p, "?")
        by_cur_h[cur][h]["spend"] += v["spend"]; by_cur_h[cur][h]["purch"] += v["purch"]
    for cur in sorted(by_cur_h):
        by_h = by_cur_h[cur]
        for h in range(24):
            v = by_h.get(h)
            if not v or (v["purch"] < 1 and v["spend"] < 1): continue
            cpa = v["spend"]/v["purch"] if v["purch"] > 0 else 0
            ws3.append([cur, f"{h:02d}시", round(v["purch"]), round(v["spend"]), round(cpa)])
        tot_s = sum(x["spend"] for x in by_h.values())
        tot_p = sum(x["purch"] for x in by_h.values())
        tot_cpa = tot_s/tot_p if tot_p > 0 else 0
        ws3.append([cur, "합계", round(tot_p), round(tot_s), round(tot_cpa)])
        for c in range(1, 6): ws3.cell(row=ws3.max_row, column=c).fill = tot_fill
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
    for r in range(2, ws3.max_row+1):
        for c in (3, 4, 5): ws3.cell(row=r, column=c).number_format = num
    style_header(ws3, 5); autowidth(ws3, [8, 9, 10, 16, 14])

    # 4·5) 피벗 매트릭스: 상품(행) × 시간대 0~23시(열), 셀 = 구매수 / 결과당비용
    #     "시간대별 구매수 테이블"과 동일 레이아웃, 셀 값만 결과당비용으로 교체한 시트도 생성.
    from openpyxl.formatting.rule import ColorScaleRule
    hours = [f"{h:02d}시" for h in range(24)]

    def build_pivot(title, kind):
        ws = wb.create_sheet(title)
        total_label = "합계" if kind == "purch" else "전체"
        ws.append(["상품", "통화"] + hours + [total_label])
        for p in prods:
            t = prod_tot[p]
            if t["purch"] < 1 and t["spend"] < 1: continue
            row = [p, prod_cur.get(p, "")]
            for h in range(24):
                v = agg.get((p, h))
                if not v or v["purch"] < 1:
                    row.append(None)                      # 구매 없으면 빈 셀(결과당비용 정의 불가)
                elif kind == "purch":
                    row.append(round(v["purch"]))
                else:
                    row.append(round(v["spend"] / v["purch"]))
            if kind == "purch":
                row.append(round(t["purch"]))
            else:
                row.append(round(t["spend"] / t["purch"]) if t["purch"] > 0 else None)
            ws.append(row)
        for r in range(2, ws.max_row + 1):
            for c in range(3, 28):                        # 00시~23시 + 합계/전체
                ws.cell(row=r, column=c).number_format = num
        style_header(ws, 27); autowidth(ws, [16, 6] + [7] * 24 + [10])
        ws.freeze_panes = "C2"                            # 상품·통화 열 + 헤더 고정
        # 행별 히트맵: 구매수는 많을수록 녹색, 결과당비용은 낮을수록 녹색
        for r in range(2, ws.max_row + 1):
            if kind == "purch":
                rule = ColorScaleRule(start_type="min", start_color="F8696B",
                                      end_type="max", end_color="63BE7B")
            else:
                rule = ColorScaleRule(start_type="min", start_color="63BE7B",
                                      end_type="max", end_color="F8696B")
            ws.conditional_formatting.add(f"C{r}:Z{r}", rule)
        return ws

    build_pivot("구매수_피벗", "purch")
    build_pivot("결과당비용_피벗", "cpr")

    fname = f"메타_시간대분석_{since}_{until}.xlsx"
    path = _desktop_dir() / fname
    try:
        wb.save(path)
    except PermissionError:
        path = _desktop_dir() / f"메타_시간대분석_{since}_{until}_{int(time.time())}.xlsx"
        wb.save(path)
    return path

def main():
    accts = [a for a in ACCOUNTS if SCOPE in ("all", a[3])]
    # (product, hour) -> {spend, purchases}; product->currency
    agg = defaultdict(lambda: {"spend": 0.0, "purch": 0.0})
    prod_cur = {}
    used = 0
    for acc, envn, cur, region in accts:
        tok = token_for(envn)
        if not tok:
            print(f"  ⏭  {acc} ({region}) — 토큰({envn}) 없음, 스킵"); continue
        print(f"  📡 {acc} ({region}, {cur}) fetch…")
        rows = fetch_insights(acc, tok)
        print(f"     rows={len(rows)}")
        used += 1
        for row in rows:
            prod = extract_product(row.get("adset_name",""), row.get("campaign_name",""))
            prod_cur.setdefault(prod, cur)
            h = hour_of(row.get("hourly_stats_aggregated_by_advertiser_time_zone"))
            sp = float(row.get("spend", 0) or 0)
            pu = purchases_of(row.get("actions"))
            agg[(prod, h)]["spend"] += sp
            agg[(prod, h)]["purch"] += pu
    if not used:
        print("\n❌ 사용 가능한 Meta 토큰이 없습니다. META_TOKEN_1 등 env 를 설정하세요.")
        return

    # 상품별 합계로 정렬
    prod_tot = defaultdict(lambda: {"spend":0.0,"purch":0.0})
    for (p,h),v in agg.items():
        prod_tot[p]["spend"]+=v["spend"]; prod_tot[p]["purch"]+=v["purch"]
    prods = sorted(prod_tot, key=lambda p:-prod_tot[p]["purch"])

    print(f"\n{'='*100}\n기간 {SINCE}~{UNTIL} · 상품별×시간대(0~23시) 판매수 / 구매당비용(CPA)\n{'='*100}")
    for p in prods:
        cur = prod_cur.get(p,"")
        t = prod_tot[p]
        if t["purch"] < 1 and t["spend"] < 1: continue
        cpa = t["spend"]/t["purch"] if t["purch"]>0 else 0
        print(f"\n■ {p}  ({cur})  — 총 판매 {t['purch']:.0f}건 · 지출 {t['spend']:,.0f} · CPA {cpa:,.0f}")
        print(f"   {'시':>3} {'판매수':>7} {'지출':>12} {'구매당비용':>12}")
        for h in range(24):
            v = agg.get((p,h))
            if not v or (v["purch"]<1 and v["spend"]<1): continue
            cpa_h = v["spend"]/v["purch"] if v["purch"]>0 else 0
            print(f"   {h:>3} {v['purch']:>7.0f} {v['spend']:>12,.0f} {cpa_h:>12,.0f}")

    # 전체 상품 합산 시간대 요약 (통화별로 분리 — 결과당비용을 정확히 보기 위해)
    print(f"\n{'='*100}\n[전체 합산] 시간대별 판매수 / 지출 / 결과당비용 (통화별)\n{'='*100}")
    # currency -> hour -> {spend, purch}
    by_cur_h = defaultdict(lambda: defaultdict(lambda: {"spend":0.0,"purch":0.0}))
    for (p,h),v in agg.items():
        cur = prod_cur.get(p, "?")
        by_cur_h[cur][h]["spend"] += v["spend"]
        by_cur_h[cur][h]["purch"] += v["purch"]
    for cur in sorted(by_cur_h):
        by_h = by_cur_h[cur]
        tot_s = sum(x["spend"] for x in by_h.values())
        tot_p = sum(x["purch"] for x in by_h.values())
        tot_cpa = tot_s/tot_p if tot_p > 0 else 0
        print(f"\n── {cur} ──  총 판매 {tot_p:.0f}건 · 지출 {tot_s:,.0f} · 결과당비용 {tot_cpa:,.0f}")
        print(f"   {'시':>3} {'판매수':>8} {'지출':>14} {'결과당비용':>12}")
        for h in range(24):
            v = by_h.get(h)
            if not v or (v["purch"] < 1 and v["spend"] < 1): continue
            cpa_h = v["spend"]/v["purch"] if v["purch"] > 0 else 0
            print(f"   {h:>3} {v['purch']:>8.0f} {v['spend']:>14,.0f} {cpa_h:>12,.0f}")

    # 바탕화면에 엑셀 저장
    try:
        path = write_excel(agg, prod_cur, prod_tot, prods, SINCE, UNTIL)
        print(f"\n✅ 엑셀 저장 완료: {path}")
    except Exception as e:
        print(f"\n❌ 엑셀 저장 실패: {e}")

if __name__ == "__main__":
    main()
